from openpyxl import load_workbook

try:
    wb = load_workbook(r'e:\LXY_learn\ETF_manager\data\数据库表导出_v3.xlsx')
    ws = wb['trading_rules(交易规则配置表)']
    print(f'行数: {ws.max_row}, 列数: {ws.max_column}')
    headers = [cell.value for cell in ws[1]]
    print(f'\n表头: {headers}')
    print(f'\n数据内容:')
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(cell is not None for cell in row):
            print(f'第{row_idx}行: {row}')
except Exception as e:
    print(e)
