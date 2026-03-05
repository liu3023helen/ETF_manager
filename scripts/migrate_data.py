"""
ETF管理系统 - 数据迁移脚本
从Excel读取13支基金数据，拆分写入SQLite各表
同时将交易规则写入trading_rules表
"""

import sqlite3
import os
import re
from datetime import date

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "data", "etf_manager.db")
EXCEL_PATH = os.path.join(BASE_DIR, "data", "基金持仓完整信息表.xlsx")
TODAY = date.today().isoformat()


def parse_price_cost(value):
    """解析 '1.032/1.123' 格式，返回 (current_nav, cost_price)"""
    if not value or "/" not in str(value):
        return None, None
    parts = str(value).split("/")
    try:
        return float(parts[0]), float(parts[1])
    except (ValueError, IndexError):
        return None, None


def parse_daily_pnl(value):
    """解析 '-19.20 (-1.15%)' 或 '+10.20 (+1.17%)' 格式，返回 (pnl, pct)"""
    if not value:
        return None, None
    s = str(value).strip()
    # 匹配 数值部分 和 百分比部分
    m = re.match(r'([+-]?\d+\.?\d*)\s*\(([+-]?\d+\.?\d*)%\)', s)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def parse_percentage(value):
    """解析 '21.37%' 或 '-2.34%' 格式"""
    if not value:
        return None
    s = str(value).strip().replace("%", "")
    try:
        return s + "%"
    except ValueError:
        return None


def classify_fund(fund_name):
    """根据基金名称自动分类"""
    name = str(fund_name)
    if any(k in name for k in ["军工", "航空航天"]):
        return "军工"
    if any(k in name for k in ["芯片", "科创"]):
        return "芯片"
    if any(k in name for k in ["黄金", "上海金"]):
        return "黄金"
    if any(k in name for k in ["机器人"]):
        return "机器人"
    if any(k in name for k in ["互联网", "港股"]):
        return "互联网"
    return "其他"


def detect_fund_type(fund_name):
    """判断基金类型"""
    name = str(fund_name)
    if "联接" in name:
        return "ETF联接"
    return "ETF"


def parse_dca_info(frequency_str):
    """解析定投频率，返回 (frequency, amount, dca_type)"""
    if not frequency_str:
        return None, None, "固定金额"

    s = str(frequency_str)
    dca_type = "固定金额"

    if "智能定投" in s:
        dca_type = "智能定投"

    # 提取频率
    freq = s
    # 提取金额
    amount_match = re.search(r'(\d+)', s)
    amount = float(amount_match.group(1)) if amount_match else None

    return freq, amount, dca_type


def migrate_excel_data(conn):
    """迁移Excel数据到SQLite"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    cursor = conn.cursor()
    fund_count = 0

    for row in range(2, ws.max_row + 1):
        platform = ws.cell(row, 1).value
        fund_name = ws.cell(row, 2).value
        fund_code = str(ws.cell(row, 3).value) if ws.cell(row, 3).value else None

        if not fund_name or not fund_code:
            continue

        # 补齐基金代码前导零（如 005693）
        fund_code = fund_code.zfill(6) if len(fund_code) < 6 else fund_code

        fund_company = ws.cell(row, 4).value
        shares = ws.cell(row, 5).value
        return_1y = ws.cell(row, 6).value
        return_3y = ws.cell(row, 7).value
        return_since = ws.cell(row, 8).value
        risk_level = ws.cell(row, 9).value
        top_holdings = ws.cell(row, 10).value
        risk_points = ws.cell(row, 11).value
        market_value = ws.cell(row, 12).value
        price_cost = ws.cell(row, 13).value
        daily_pnl_str = ws.cell(row, 14).value
        position_pct = ws.cell(row, 15).value
        total_dca = ws.cell(row, 16).value
        dca_freq = ws.cell(row, 17).value
        is_dca = ws.cell(row, 18).value

        fund_category = classify_fund(fund_name)
        fund_type = detect_fund_type(fund_name)

        # ===== 写入 fund_info =====
        cursor.execute("""
            INSERT OR REPLACE INTO fund_info
            (fund_code, fund_name, fund_company, fund_type, risk_level,
             fund_category, top_holdings, risk_points,
             return_1y, return_3y, return_since_inception)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fund_code, fund_name, fund_company, fund_type, risk_level,
            fund_category, top_holdings, risk_points,
            str(return_1y) if return_1y else None,
            str(return_3y) if return_3y else None,
            str(return_since) if return_since else None,
        ))

        # ===== 写入 my_holdings =====
        current_nav, cost_price = parse_price_cost(price_cost)
        shares_val = float(shares) if shares else 0
        base_shares = shares_val * 0.2
        tradable_shares = shares_val - base_shares
        total_invested_val = float(total_dca) if total_dca else (float(market_value) if market_value else 0)

        cursor.execute("""
            INSERT OR REPLACE INTO my_holdings
            (fund_code, platform, shares, cost_price, base_shares, tradable_shares,
             current_value, total_invested, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fund_code, platform, shares_val, cost_price or 0,
            base_shares, tradable_shares,
            float(market_value) if market_value else 0,
            total_invested_val, TODAY,
        ))

        # ===== 写入 daily_quotes（涨乐财富通有当日数据）=====
        if daily_pnl_str:
            pnl, pct = parse_daily_pnl(daily_pnl_str)
            cursor.execute("""
                INSERT OR REPLACE INTO daily_quotes
                (fund_code, date, nav, daily_change_pct, daily_value, daily_pnl)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                fund_code, TODAY, current_nav, pct,
                float(market_value) if market_value else None,
                pnl,
            ))

        # ===== 写入 dca_plans（支付宝有定投信息）=====
        if is_dca == "是" or dca_freq:
            freq, amount, dca_type = parse_dca_info(dca_freq)
            cursor.execute("""
                INSERT OR REPLACE INTO dca_plans
                (fund_code, platform, is_active, frequency, amount, dca_type, total_invested)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                fund_code, platform, 1 if is_dca == "是" else 0,
                freq, amount, dca_type,
                float(total_dca) if total_dca else 0,
            ))

        fund_count += 1
        print(f"  已迁移: [{platform}] {fund_name} ({fund_code}) - {fund_category}")

    conn.commit()
    wb.close()
    return fund_count


def insert_trading_rules(conn):
    """将交易规则写入trading_rules表"""
    cursor = conn.cursor()

    # 先清空旧规则
    cursor.execute("DELETE FROM trading_rules;")

    rules = [
        # 军工/芯片止盈规则
        ("军工", "止盈", "从60日低点涨≥30% 或 单月涨≥15%", 0.30, "卖出一半（两条件不叠加）", 1, 1),
        ("芯片", "止盈", "从60日低点涨≥30% 或 单月涨≥15%", 0.30, "卖出一半（两条件不叠加）", 1, 1),
        ("军工", "止盈", "从60日低点涨≥50%", 0.50, "再卖剩余的一半，保留1/5底仓", 2, 1),
        ("芯片", "止盈", "从60日低点涨≥50%", 0.50, "再卖剩余的一半，保留1/5底仓", 2, 1),

        # 黄金止盈规则
        ("黄金", "止盈", "涨≥10%", 0.10, "减仓1/3，保留1/5底仓", 1, 1),

        # 军工/芯片接回规则
        ("军工", "接回", "从卖出价回落10%", -0.10, "买回卖出量的1/3", 1, 1),
        ("芯片", "接回", "从卖出价回落10%", -0.10, "买回卖出量的1/3", 1, 1),
        ("军工", "接回", "从卖出价回落20%", -0.20, "再买回卖出量的1/3", 2, 1),
        ("芯片", "接回", "从卖出价回落20%", -0.20, "再买回卖出量的1/3", 2, 1),
        ("军工", "接回", "从卖出价回落30%", -0.30, "买回剩余，恢复原仓位", 3, 1),
        ("芯片", "接回", "从卖出价回落30%", -0.30, "买回剩余，恢复原仓位", 3, 1),

        # 黄金接回规则
        ("黄金", "接回", "从卖出价回落5%-8%", -0.05, "分4周定投买回，每周买回减仓量的1/4", 1, 1),

        # 大跌加仓规则
        ("军工", "加仓", "当日跌2%-3%", -0.02, "+100元（军工/芯片任选）", 1, 1),
        ("芯片", "加仓", "当日跌2%-3%", -0.02, "+100元（军工/芯片任选）", 1, 1),
        ("军工", "加仓", "当日跌3%-5%", -0.03, "+200元（军工/芯片任选）", 2, 1),
        ("芯片", "加仓", "当日跌3%-5%", -0.03, "+200元（军工/芯片任选）", 2, 1),
        ("军工", "加仓", "当日跌>5%", -0.05, "+300元（军工/芯片任选）", 3, 1),
        ("芯片", "加仓", "当日跌>5%", -0.05, "+300元（军工/芯片任选）", 3, 1),

        # 核心纪律
        ("全部", "纪律", "任何时候保留1/5仓位不卖", 0.20, "底仓不动", 0, 1),
        ("全部", "纪律", "定投永不停止", 0, "无论涨跌，定投继续", 0, 1),
        ("全部", "纪律", "大涨当天不额外买入", 0, "不追涨", 0, 1),
    ]

    cursor.executemany("""
        INSERT INTO trading_rules
        (fund_category, rule_type, condition_desc, threshold, action_desc, priority, is_active)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, rules)

    conn.commit()
    print(f"\n  已写入 {len(rules)} 条交易规则")
    return len(rules)


def verify_migration(conn):
    """验证迁移结果"""
    cursor = conn.cursor()
    print("\n===== 迁移验证 =====")

    tables = ["fund_info", "my_holdings", "daily_quotes", "dca_plans", "trading_rules"]
    for table in tables:
        cursor.execute(f"SELECT COUNT(*) FROM {table};")
        count = cursor.fetchone()[0]
        print(f"  {table:20s}: {count} 条记录")

    # 检查外键完整性
    cursor.execute("""
        SELECT h.fund_code, h.platform
        FROM my_holdings h
        LEFT JOIN fund_info f ON h.fund_code = f.fund_code
        WHERE f.fund_code IS NULL
    """)
    orphans = cursor.fetchall()
    if orphans:
        print(f"\n  [警告] 发现 {len(orphans)} 条孤立持仓记录!")
    else:
        print("\n  外键完整性检查: 通过")

    # 显示基金分类统计
    cursor.execute("SELECT fund_category, COUNT(*) FROM fund_info GROUP BY fund_category;")
    print("\n  基金分类统计:")
    for cat, cnt in cursor.fetchall():
        print(f"    {cat}: {cnt}支")

    # 显示交易规则统计
    cursor.execute("SELECT rule_type, COUNT(*) FROM trading_rules GROUP BY rule_type;")
    print("\n  交易规则统计:")
    for rtype, cnt in cursor.fetchall():
        print(f"    {rtype}: {cnt}条")


def main():
    if not os.path.exists(DB_PATH):
        print(f"错误: 数据库不存在 {DB_PATH}")
        print("请先运行 init_db.py")
        return

    if not os.path.exists(EXCEL_PATH):
        print(f"错误: Excel文件不存在 {EXCEL_PATH}")
        return

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys=ON;")

    print("===== 开始数据迁移 =====\n")

    print("1. 迁移Excel基金数据...")
    fund_count = migrate_excel_data(conn)
    print(f"\n  共迁移 {fund_count} 支基金")

    print("\n2. 写入交易规则...")
    rule_count = insert_trading_rules(conn)

    verify_migration(conn)

    conn.close()
    print("\n数据迁移完成!")


if __name__ == "__main__":
    main()
