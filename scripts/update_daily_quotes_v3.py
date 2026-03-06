from openpyxl import load_workbook
from datetime import datetime

# 准备数据
# 格式: (fund_code, fund_name, quote_date, nav, daily_change_pct, daily_pnl, daily_value)
quotes_data = [
    # 场内ETF (日期: 2026-03-05) - 涨跌幅未知，当日盈亏未知(总计+93.90)
    ('159530', '机器人ETF易方达', '2026-03-05', 1.464, None, None, 2196.00),
    ('159770', '机器人ETF', '2026-03-05', 1.055, None, None, 1688.00),
    ('159781', '科创创业ETF易方达', '2026-03-05', 0.922, None, None, 1383.00),
    ('159241', '航空航天ETF天弘', '2026-03-05', 1.511, None, None, 906.60),
    ('159227', '航空航天ETF', '2026-03-05', 1.483, None, None, 889.80),
    ('159792', '港股通互联网ETF', '2026-03-05', 0.717, None, None, 143.40),

    # 场外ETF联接 (日期: 2026-03-05) - 根据昨日净值和今日涨跌幅估算
    # NAV估算 = 昨净值 * (1 + 涨幅%)
    # 005693: 1.4273 * 1.0129 = 1.4457
    ('005693', '广发中证军工ETF联接C', '2026-03-05', 1.4457, 1.29, 4.74, 471.25),
    # 017470: 2.0598 * 0.9957 = 2.0509
    ('017470', '嘉实上证科创板芯片ETF联接C', '2026-03-05', 2.0509, -0.43, -3.86, 893.79),
    # 018345: 1.2061 * 0.9945 = 1.1995
    ('018345', '华夏中证机器人ETF联接C', '2026-03-05', 1.1995, -0.55, -10.35, 1863.57),
    # 020482: 1.5400 * 0.9945 = 1.5315
    ('020482', '招商中证机器人ETF联接C', '2026-03-05', 1.5315, -0.55, -10.28, 1863.17),
    # 014881: 1.1846 * 0.9945 = 1.1781
    ('014881', '天弘中证机器人ETF联接C', '2026-03-05', 1.1781, -0.55, -10.38, 1862.87),

    # 黄金ETF联接 (日期: 2026-03-04) - 数据明确为03-04
    ('014662', '天弘上海金ETF联接C', '2026-03-04', 2.5672, -0.24, -15.51, 598.57),
    ('009034', '建信上海金ETF联接C', '2026-03-04', 2.6560, 4.14, -26.31, 1020.76),
    ('008702', '华夏黄金ETF联接C', '2026-03-04', 2.4976, -0.19, -30.60, 1197.75),
    ('009505', '富国上海金ETF联接C', '2026-03-04', 2.4192, 0.80, -20.81, 1007.96),
    ('020341', '工银瑞信黄金ETF联接E', '2026-03-04', 2.6159, 1.98, -20.74, 815.82),

    # 货币基金 (日期: 2026-03-05)
    ('002183', '广发天天红货币B', '2026-03-05', 1.0000, 0, 3.00, 81064.63)
]

file_path = r'e:\LXY_learn\ETF_manager\data\数据库表导出_v3_updated.xlsx'

try:
    wb = load_workbook(file_path)
    sheet_name = 'daily_quotes(每日净值表)'
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 获取现有的最大ID
        max_id = 0
        if ws.max_row > 1:
            # 假设第一列是ID
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 写入新数据
        for item in quotes_data:
            max_id += 1
            # 字段顺序: quote_id, fund_code, fund_name, quote_date, nav, acc_nav, daily_change_pct, daily_value, daily_pnl, created_at
            ws.append([
                max_id,             # quote_id
                item[0],            # fund_code
                item[1],            # fund_name
                item[2],            # quote_date
                item[3],            # nav
                '',                 # acc_nav (空)
                item[4],            # daily_change_pct
                item[6],            # daily_value
                item[5],            # daily_pnl
                current_time        # created_at
            ])
            
        wb.save(file_path)
        print(f"成功添加 {len(quotes_data)} 条净值记录到 {sheet_name}")
    else:
        print(f"错误：找不到工作表 {sheet_name}")

except Exception as e:
    print(f"发生错误: {e}")
