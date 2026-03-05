"""
ETF管理系统 - Excel报表导出脚本
从SQLite查询数据，生成多Sheet的Excel报表
"""

import sqlite3
import sys
import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import DB_PATH, OUTPUT_DIR

OUTPUT_PATH = os.path.join(OUTPUT_DIR, "ETF持仓报表.xlsx")


def style_header(ws, row, col_count):
    """设置表头样式"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, col_count + 1):
        cell = ws.cell(row, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, col_count):
    """设置数据行样式"""
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row, col)
            cell.alignment = data_align
            cell.border = thin_border
            if (row - start_row) % 2 == 1:
                cell.fill = alt_fill


def export_holdings_sheet(wb, conn):
    """导出持仓总览Sheet"""
    ws = wb.active
    ws.title = "持仓总览"

    headers = ["平台", "基金名称", "基金代码", "基金公司", "分类", "风险等级",
               "持仓份额", "成本价", "当前市值", "底仓份额", "可交易份额",
               "累计投入", "近1年收益率", "近3年收益率"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers))

    cursor = conn.cursor()
    cursor.execute("""
        SELECT h.platform, f.fund_name, f.fund_code, f.fund_company,
               f.fund_category, f.risk_level,
               h.shares, h.cost_price, h.current_value,
               h.base_shares, h.tradable_shares, h.total_invested,
               f.return_1y, f.return_3y
        FROM my_holdings h
        JOIN fund_info f ON h.fund_code = f.fund_code
        ORDER BY h.platform, f.fund_category;
    """)

    row = 2
    for data in cursor.fetchall():
        for col, val in enumerate(data, 1):
            ws.cell(row, col, val)
        row += 1

    style_data_rows(ws, 2, row - 1, len(headers))

    # 设置列宽
    widths = [14, 28, 10, 14, 8, 12, 10, 10, 10, 10, 10, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return row - 2


def export_dca_sheet(wb, conn):
    """导出定投计划Sheet"""
    ws = wb.create_sheet("定投计划")

    headers = ["基金名称", "基金代码", "平台", "状态", "定投频率",
               "每次金额", "定投类型", "累计定投金额"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers))

    cursor = conn.cursor()
    cursor.execute("""
        SELECT f.fund_name, f.fund_code, d.platform,
               CASE WHEN d.is_active = 1 THEN '启用' ELSE '停用' END,
               d.frequency, d.amount, d.dca_type, d.total_invested
        FROM dca_plans d
        JOIN fund_info f ON d.fund_code = f.fund_code
        ORDER BY d.platform;
    """)

    row = 2
    for data in cursor.fetchall():
        for col, val in enumerate(data, 1):
            ws.cell(row, col, val)
        row += 1

    style_data_rows(ws, 2, row - 1, len(headers))

    widths = [28, 10, 14, 8, 28, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return row - 2


def export_rules_sheet(wb, conn):
    """导出交易规则Sheet"""
    ws = wb.create_sheet("交易规则")

    headers = ["基金分类", "规则类型", "触发条件", "阈值", "操作", "优先级", "状态"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers))

    cursor = conn.cursor()
    cursor.execute("""
        SELECT fund_category, rule_type, condition_desc, threshold,
               action_desc, priority,
               CASE WHEN is_active = 1 THEN '启用' ELSE '停用' END
        FROM trading_rules
        ORDER BY fund_category, rule_type, priority;
    """)

    row = 2
    for data in cursor.fetchall():
        for col, val in enumerate(data, 1):
            ws.cell(row, col, val)
        row += 1

    style_data_rows(ws, 2, row - 1, len(headers))

    widths = [12, 10, 35, 10, 40, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return row - 2


def export_fund_info_sheet(wb, conn):
    """导出基金详情Sheet"""
    ws = wb.create_sheet("基金详情")

    headers = ["基金代码", "基金名称", "基金公司", "类型", "分类",
               "风险等级", "前三大重仓股及占比", "主要风险点"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers))

    cursor = conn.cursor()
    cursor.execute("""
        SELECT fund_code, fund_name, fund_company, fund_type,
               fund_category, risk_level, top_holdings, risk_points
        FROM fund_info
        ORDER BY fund_category, fund_code;
    """)

    row = 2
    for data in cursor.fetchall():
        for col, val in enumerate(data, 1):
            ws.cell(row, col, val)
        row += 1

    style_data_rows(ws, 2, row - 1, len(headers))

    widths = [10, 28, 14, 10, 8, 12, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return row - 2


def main():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    conn = sqlite3.connect(DB_PATH)

    wb = openpyxl.Workbook()

    print("导出Excel报表...\n")

    cnt1 = export_holdings_sheet(wb, conn)
    print(f"  持仓总览: {cnt1} 条")

    cnt2 = export_dca_sheet(wb, conn)
    print(f"  定投计划: {cnt2} 条")

    cnt3 = export_rules_sheet(wb, conn)
    print(f"  交易规则: {cnt3} 条")

    cnt4 = export_fund_info_sheet(wb, conn)
    print(f"  基金详情: {cnt4} 条")

    wb.save(OUTPUT_PATH)
    conn.close()

    print(f"\n报表已导出: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
